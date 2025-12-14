import logging
import tempfile
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from pyxlsb import open_workbook as open_xlsb
from pathlib import Path
from typing import Dict, Any, List, Optional

logger = logging.getLogger(__name__)

IMAGES_DIR = Path("./uploads/shoe_images")
IMAGES_DIR.mkdir(parents=True, exist_ok=True)


def detect_style_column(sheet, start_row: int, max_col: int) -> Optional[int]:
    """Detect style column by sampling: 6 digits only, no letters"""
    logger.info("🔍 Detecting STYLE column by sampling data...")
    
    for col_idx in range(1, min(max_col + 1, 10)):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        sample_rows = [start_row + i for i in range(1, min(11, sheet.max_row - start_row + 1))]
        valid_count = 0
        
        for row_idx in sample_rows:
            val = sheet[f"{col_letter}{row_idx}"].value
            if val:
                val_str = str(val).strip()
                if val_str.isdigit() and len(val_str) == 6:
                    valid_count += 1
        
        if valid_count >= 7:
            logger.info(f"  ✅ STYLE column detected: {col_letter} (column {col_idx}) - {valid_count}/10 samples matched")
            return col_idx
    
    return None


def detect_color_column(sheet, start_row: int, max_col: int) -> Optional[int]:
    """Detect color column: header must be exactly 'Color' (case-insensitive), data is 3-4 letters"""
    logger.info("🔍 Detecting COLOR column...")
    
    for col_idx in range(1, min(max_col + 1, 20)):
        header_val = sheet.cell(row=1, column=col_idx).value
        if header_val and str(header_val).lower().strip() == 'color':
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            logger.info(f"  ✅ COLOR column detected by header: {col_letter} (column {col_idx})")
            return col_idx
    
    for col_idx in range(1, min(max_col + 1, 10)):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        sample_rows = [start_row + i for i in range(1, min(11, sheet.max_row - start_row + 1))]
        valid_count = 0
        
        for row_idx in sample_rows:
            val = sheet[f"{col_letter}{row_idx}"].value
            if val:
                val_str = str(val).strip().upper()
                
                if val_str in ['MENS', 'WOMENS', 'WOMEN', 'KIDS', 'BOYS', 'GIRLS', 'UNISEX', 'MALE', 'FEMALE', 'MEN', 'YOUTH']:
                    continue
                
                if val_str.isalpha() and 3 <= len(val_str) <= 4:
                    valid_count += 1
        
        if valid_count >= 7:
            logger.info(f"  ✅ COLOR column detected by pattern: {col_letter} (column {col_idx}) - {valid_count}/10 samples matched")
            return col_idx
    
    logger.info("  ⚠️  No dedicated COLOR column found")
    return None


def detect_image_column(sheet, start_row: int, max_col: int) -> Optional[int]:
    """Detect image column - usually column A (1) with header 'Image'"""
    logger.info("🔍 Detecting IMAGE column...")
    
    for col_idx in range(1, min(max_col + 1, 5)):
        val = sheet.cell(row=1, column=col_idx).value
        if val and 'image' in str(val).lower():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            logger.info(f"  ✅ IMAGE column detected: {col_letter} (column {col_idx}) from header")
            return col_idx
    
    logger.info(f"  ⚠️  IMAGE column not found in headers, defaulting to column A")
    return 1


def detect_header_column(sheet, header_row: int, column_name: str) -> Optional[int]:
    """Detect column by exact header name match"""
    logger.info(f"🔍 Looking for '{column_name}' column in header...")
    
    row = [cell.value for cell in sheet[header_row]]
    for col_idx, cell_value in enumerate(row, 1):
        if cell_value:
            val_lower = str(cell_value).lower().strip()
            if column_name.lower() in val_lower:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                logger.info(f"  ✅ {column_name.upper()} column found: {col_letter} (column {col_idx})")
                return col_idx
    
    logger.info(f"  ⚠️  {column_name.upper()} column not found")
    return None


def detect_sku_column(sheet, header_row: int, max_col: int) -> Optional[int]:
    """Detect SKU column - contains style_color format like '104289_WSL'"""
    logger.info("🔍 Detecting SKU column...")
    
    for col_idx in range(1, min(max_col + 1, 10)):
        header_val = sheet.cell(row=header_row, column=col_idx).value
        if header_val and 'sku' in str(header_val).lower():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            logger.info(f"  ✅ SKU column detected by header: {col_letter} (column {col_idx})")
            return col_idx
    
    logger.info("  ⚠️  SKU column not found")
    return None


def extract_color_from_sku(sku: str) -> Optional[str]:
    """Extract color code from SKU format like '104289_WSL' -> 'WSL'"""
    if not sku:
        return None
    
    sku_str = str(sku).strip()
    if '_' in sku_str:
        parts = sku_str.split('_')
        if len(parts) >= 2:
            color = parts[-1].strip()
            if color.isalpha() and 2 <= len(color) <= 5:
                return color
    
    return None


def extract_and_save_image(image_loader, cell_address: str, style: str, color: str) -> Optional[str]:
    """Extract image from cell and save with style_color naming"""
    if not image_loader:
        return None
    
    try:
        img = image_loader.get(cell_address)
        if img:
            safe_style = ''.join(c for c in style if c.isalnum() or c in '-_')
            safe_color = ''.join(c for c in color if c.isalnum() or c in '-_')
            
            if not safe_style or not safe_color:
                logger.warning(f"  ⚠️  Invalid filename components for {style}_{color}")
                return None
            
            filename = f"{safe_style}_{safe_color}.png"
            filepath = IMAGES_DIR / filename
            
            img.save(str(filepath), 'PNG')
            
            image_url = f"/uploads/shoe_images/{filename}"
            logger.info(f"  💾 Saved image: {filename} -> {image_url}")
            return image_url
        else:
            return None
    except KeyError:
        return None
    except Exception as e:
        logger.warning(f"  ⚠️  Could not extract image for {style}_{color} at {cell_address}: {e}")
    
    return None


def parse_excel_ki(file_path: Path) -> List[Dict]:
    """Parse KI sheet - extract style, color, image, division, outsole"""
    logger.info(f"📊 Parsing KI sheet: {file_path.name}")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    if sheet is None:
        logger.warning("⚠️  Active sheet is None, trying first sheet...")
        if wb.worksheets:
            sheet = wb.worksheets[0]
            logger.info(f"✅ Using first sheet: {sheet.title}")
        else:
            raise ValueError("File appears to be corrupted (no worksheets found).")
    
    logger.info(f"📄 Sheet name: {sheet.title}, Max row: {sheet.max_row}, Max col: {sheet.max_column}")
    
    header_row = 1
    for row_idx in range(1, 11):
        row = [cell.value for cell in sheet[row_idx]]
        row_str = ' '.join([str(v).lower() if v else '' for v in row])
        if 'style' in row_str or 'color' in row_str:
            header_row = row_idx
            logger.info(f"📍 Found header row at {row_idx}")
            break
    
    col_map = {}
    data_start_row = header_row + 1
    
    col_map['style'] = detect_style_column(sheet, data_start_row, sheet.max_column)
    col_map['color'] = detect_color_column(sheet, data_start_row, sheet.max_column)
    col_map['sku'] = detect_sku_column(sheet, header_row, sheet.max_column)
    
    col_map['image'] = detect_image_column(sheet, header_row, sheet.max_column)
    col_map['division'] = detect_header_column(sheet, header_row, 'division')
    col_map['outsole'] = detect_header_column(sheet, header_row, 'outsole')
    col_map['colorDescription'] = detect_header_column(sheet, header_row, 'color description')
    
    if not col_map['style']:
        raise ValueError(f"Could not detect style column. Found: {col_map}")
    
    if not col_map['color'] and not col_map['sku']:
        raise ValueError(f"Could not detect color or SKU column. Found: {col_map}")
    
    if not col_map['color'] and col_map['sku']:
        logger.info("  ℹ️  No dedicated color column - will extract from SKU")
    
    image_loader = None
    if col_map.get('image'):
        try:
            image_loader = SheetImageLoader(sheet)
            logger.info("✅ Image loader initialized")
        except Exception as e:
            logger.warning(f"⚠️  Could not initialize image loader: {e}")
            image_loader = None
    
    color_info = openpyxl.utils.get_column_letter(col_map['color']) if col_map['color'] else f"SKU({openpyxl.utils.get_column_letter(col_map['sku'])})" if col_map['sku'] else 'N/A'
    logger.info(f"✅ Column mapping: Style={openpyxl.utils.get_column_letter(col_map['style'])}, Color={color_info}, Image={openpyxl.utils.get_column_letter(col_map['image']) if col_map['image'] else 'N/A'}, Division={openpyxl.utils.get_column_letter(col_map['division']) if col_map['division'] else 'N/A'}, Outsole={openpyxl.utils.get_column_letter(col_map['outsole']) if col_map['outsole'] else 'N/A'}")
    
    items = []
    skipped = 0
    images_extracted = 0
    
    for row_idx in range(data_start_row, sheet.max_row + 1):
        row = sheet[row_idx]
        
        style_val = row[col_map['style'] - 1].value
        
        if col_map['color']:
            color_val = row[col_map['color'] - 1].value
        elif col_map['sku']:
            sku_val = row[col_map['sku'] - 1].value
            color_val = extract_color_from_sku(sku_val)
        else:
            color_val = None
        
        if row_idx <= header_row + 3:
            logger.info(f"Row {row_idx}: Style='{style_val}', Color='{color_val}'")
        
        if not style_val or not color_val:
            skipped += 1
            continue
        
        style = str(style_val).strip()
        color = str(color_val).strip()
        
        if not style or not color or style.lower() == 'none' or color.lower() == 'none':
            skipped += 1
            continue
        
        image_url = None
        if image_loader and col_map.get('image'):
            col_letter = openpyxl.utils.get_column_letter(col_map['image'])
            cell_address = f"{col_letter}{row_idx}"
            image_url = extract_and_save_image(image_loader, cell_address, style, color)
            if image_url:
                images_extracted += 1
        
        item = {
            "style": style,
            "color": color,
            "image": image_url
        }
        
        if col_map.get('division'):
            division_val = row[col_map['division'] - 1].value
            if division_val:
                item["division"] = str(division_val).strip()
        
        if col_map.get('outsole'):
            outsole_val = row[col_map['outsole'] - 1].value
            if outsole_val:
                item["outsole"] = str(outsole_val).strip()
        
        if col_map.get('colorDescription'):
            color_desc_val = row[col_map['colorDescription'] - 1].value
            if color_desc_val:
                item["colorDescription"] = str(color_desc_val).strip()
        
        items.append(item)
    
    wb.close()
    logger.info(f"✅ Parsed {len(items)} KI items (skipped {skipped} empty rows)")
    logger.info(f"🖼️  Extracted {images_extracted} images")
    
    if items:
        logger.info(f"Sample items: {items[:3]}")
    
    return items


def parse_excel_allbought(file_path: Path) -> List[Dict]:
    """Parse All Bought sheet - extract style, color, image, division, outsole"""
    logger.info(f"📊 Parsing All Bought sheet: {file_path.name}")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    if sheet is None:
        logger.warning("⚠️  Active sheet is None, trying first sheet...")
        if wb.worksheets:
            sheet = wb.worksheets[0]
            logger.info(f"✅ Using first sheet: {sheet.title}")
        else:
            raise ValueError("File appears to be corrupted (no worksheets found).")
    
    header_row = None
    col_map = {}
    
    for row_idx in range(1, 11):
        row = [cell.value for cell in sheet[row_idx]]
        row_str = ' '.join([str(v).lower() if v else '' for v in row])
        
        if 'style' in row_str and 'color' in row_str:
            header_row = row_idx
            for col_idx, cell_value in enumerate(row, 1):
                if cell_value:
                    val_lower = str(cell_value).lower().strip()
                    if 'style' in val_lower:
                        col_map['style'] = col_idx
                    elif 'color description' in val_lower or 'colordescription' in val_lower:
                        col_map['colorDescription'] = col_idx
                    elif val_lower == 'color':
                        col_map['color'] = col_idx
                    elif 'image' in val_lower or 'photo' in val_lower or 'picture' in val_lower:
                        col_map['image'] = col_idx
                    elif 'division' in val_lower:
                        col_map['division'] = col_idx
                    elif 'outsole' in val_lower:
                        col_map['outsole'] = col_idx
            break
    
    if not header_row or 'style' not in col_map or 'color' not in col_map:
        raise ValueError("Could not find style and color columns")
    
    logger.info(f"✓ Header at row {header_row}, columns: {list(col_map.keys())}")
    
    items = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        row = sheet[row_idx]
        
        style_val = row[col_map['style'] - 1].value
        color_val = row[col_map['color'] - 1].value
        
        if not style_val or not color_val:
            continue
        
        style = str(style_val).strip()
        color = str(color_val).strip()
        
        if not style or not color:
            continue
        
        item = {
            "style": style,
            "color": color
        }
        
        if 'image' in col_map:
            image_val = row[col_map['image'] - 1].value
            if image_val:
                image_str = str(image_val).strip()
                import re
                if '/images/' in image_str:
                    match = re.search(r'/images/([^/\s]+)$', image_str)
                    if match:
                        filename = match.group(1)
                        image_str = f"/uploads/shoe_images/{filename}"
                item["image"] = image_str
        
        if 'division' in col_map:
            division_val = row[col_map['division'] - 1].value
            if division_val:
                item["division"] = str(division_val).strip()
        
        if 'outsole' in col_map:
            outsole_val = row[col_map['outsole'] - 1].value
            if outsole_val:
                item["outsole"] = str(outsole_val).strip()
        
        if 'colorDescription' in col_map:
            color_desc_val = row[col_map['colorDescription'] - 1].value
            if color_desc_val:
                item["colorDescription"] = str(color_desc_val).strip()
        
        items.append(item)
    
    wb.close()
    logger.info(f"✅ Parsed {len(items)} All Bought items")
    return items


def parse_xlsb_file(file_path: Path) -> List[Dict]:
    """Parse XLSB (binary Excel) file - works for both KI and AllBought"""
    logger.info(f"📊 Parsing XLSB file: {file_path.name}")
    
    items = []
    
    with open_xlsb(str(file_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            raise ValueError("No sheets found in XLSB file")
        
        sheet_name = sheet_names[0]
        logger.info(f"📄 Reading sheet: {sheet_name}")
        
        with wb.get_sheet(sheet_name) as sheet:
            rows = list(sheet.rows())
            
            if not rows:
                raise ValueError("Sheet is empty")
            
            header_row_idx = None
            col_map = {}
            
            for idx, row in enumerate(rows[:10]):
                row_values = [cell.v if cell else None for cell in row]
                row_str = ' '.join([str(v).lower() if v else '' for v in row_values])
                
                if 'style' in row_str and 'color' in row_str:
                    header_row_idx = idx
                    for col_idx, cell in enumerate(row):
                        if cell and cell.v:
                            val_lower = str(cell.v).lower().strip()
                            if 'style' in val_lower:
                                col_map['style'] = col_idx
                            elif 'color description' in val_lower or 'colordescription' in val_lower:
                                col_map['colorDescription'] = col_idx
                            elif val_lower == 'color':
                                col_map['color'] = col_idx
                    break
            
            if not header_row_idx or 'style' not in col_map or 'color' not in col_map:
                raise ValueError("Could not find style and color columns in XLSB")
            
            logger.info(f"✓ Header at row {header_row_idx}, Style col: {col_map['style']}, Color col: {col_map['color']}")
            
            for row in rows[header_row_idx + 1:]:
                row_values = [cell.v if cell else None for cell in row]
                
                style = row_values[col_map['style']] if col_map['style'] < len(row_values) else None
                color = row_values[col_map['color']] if col_map['color'] < len(row_values) else None
                
                if not style or not color:
                    continue
                
                style_str = str(style).strip()
                color_str = str(color).strip()
                
                if not style_str or not color_str:
                    continue
                
                item = {
                    "style": style_str,
                    "color": color_str,
                    "image": None
                }
                
                if 'colorDescription' in col_map and col_map['colorDescription'] < len(row_values):
                    color_desc = row_values[col_map['colorDescription']]
                    if color_desc:
                        item["colorDescription"] = str(color_desc).strip()
                
                items.append(item)
    
    logger.info(f"✅ Parsed {len(items)} items from XLSB")
    return items
